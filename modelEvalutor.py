import pymssql
import os
import numpy as np
import torch
from torchvision import models, transforms
from PIL import Image
import sqlite3  # Import SQLite library
from scipy.spatial.distance import cosine
import gc  # Garbage Collector interface
import openpyxl
import struct

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ViT16"
ws.cell(row=1, column=1, value='Species')
ws.cell(row=1, column=2, value='Total Available Data')
ws.cell(row=1, column=3, value='Total Data Used')
ws.cell(row=1, column=4, value='Top 1 Score Success')
ws.cell(row=1, column=5, value='Top 10 Score Success')
ws.cell(row=1, column=6, value='Average Success Score')

connection = pymssql.connect(
            server="128.46.81.40",
            user="forbeslab",
            password="r^%o*3zMzsgzSxADi9",
            database="Fathom"
)
cursor = connection.cursor()

cursor.execute("SELECT id, concept FROM bounding_boxes ORDER BY id")
concept_to_ids_map = {}
id_to_concept_map = {}
total_ids = 0
float_size = 4 
for row in cursor.fetchall():
    total_ids+=1
    concept = row[1]
    id = row[0]
    id_to_concept_map[id]=concept
    if concept not in concept_to_ids_map:
        concept_to_ids_map[concept] = []
    concept_to_ids_map[concept].append(id)

concept_list = list(concept_to_ids_map.keys())
concept_to_cell_map = {}
for i in range(len(concept_list)):
    concept_to_cell_map[concept_list[i]] = i+2
    ws.cell(row=i + 2, column=1, value=concept_list[i])
    ws.cell(row=i + 2, column=2, value=len(concept_to_ids_map[concept_list[i]]))
    ws.cell(row=i + 2, column=3, value=0)
    ws.cell(row=i + 2, column=4, value=0)
    ws.cell(row=i + 2, column=5, value=0)
    ws.cell(row=i + 2, column=6, value=0)

file_path = './similarity_scores.bin'

with open(file_path, 'rb') as file:
    for i in range(120000):
        if i % 1000 == 0:
            print(i)
        file.seek((total_ids+1) * (i+1) * float_size)
            
        data = file.read((total_ids+1) * float_size)
            
        float_array = struct.unpack(f'{(total_ids+1)}f', data)
        if(float_array[i+1]!=0.0):
            curCellIndex = concept_to_cell_map[id_to_concept_map[int(float_array[0])]]

            ws.cell(row=curCellIndex, column=3, value=int(ws.cell(row=curCellIndex, column=3).value)+1)

            top_3_tuples = sorted(enumerate(float_array), key=lambda x: x[1], reverse=True)[:13]

            for j in range(10):
                file.seek((top_3_tuples[2+j][0]) * float_size)
                id_data = file.read(float_size)
                id_num = int(struct.unpack(f'1f', id_data)[0])

                curTop1SuccessCount = int(ws.cell(row=curCellIndex, column=4).value)
                curTop10SuccessCount = int(ws.cell(row=curCellIndex, column=5).value)


                if(id_to_concept_map[id_num] == id_to_concept_map[top_3_tuples[0][1]]):#success
                    if j==0:
                        ws.cell(row=curCellIndex, column=4, value=curTop1SuccessCount+1)
                        ws.cell(row=curCellIndex, column=6, value=(float(ws.cell(row=curCellIndex, column=6).value) * curTop1SuccessCount + top_3_tuples[2+j][1])/(curTop1SuccessCount+1))
                    ws.cell(row=curCellIndex, column=5, value=curTop10SuccessCount+1)
                    break
            #else:
                #error
            #    ws.cell(row=curCellIndex, column=5, value=(int(ws.cell(row=curCellIndex, column=4).value) * curDataCount + top_3_tuples[2][1])/(curDataCount+1))



wb.save("./scoreResult.xlsx")