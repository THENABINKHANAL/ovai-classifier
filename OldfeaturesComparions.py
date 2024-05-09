import pymssql
import os
import numpy as np
import torch
from torchvision import models, transforms
from PIL import Image, UnidentifiedImageError
import openpyxl
from scipy.spatial.distance import cosine
import gc  # Garbage Collector interface

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Get the active worksheet
ws = wb.active

# Rename the sheet
ws.title = "MySheet"

FOLDER_DIRECTORY = "./images"
ws['A1'] = 'ID'

model = models.efficientnet_b7(pretrained=True)
model.eval()  # Set the model to evaluation mode

preprocess = transforms.Compose([
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
])

BATCH_SIZE = 4000  # Adjusted batch size for database queries
IMAGE_BATCH_SIZE = 4000  # Adjusted batch size for image processing to manage memory usage better

def batch_extract_features(image_paths, bboxes):
    batch_tensors = []
    feature_vectors = [None] * len(image_paths)

    for img_path, bbox in zip(image_paths, bboxes):
        try:
            image = Image.open(img_path).convert('RGB')
            roi = image.crop(bbox)
            tensor = preprocess(roi).unsqueeze(0)
            batch_tensors.append(tensor)
        except:
            print(f"Cannot identify image file {img_path}")
            batch_tensors.append(None)

    
    if not any(t is not None for t in batch_tensors):
        return feature_vectors

    valid_tensors = torch.cat([t for t in batch_tensors if t is not None], 0)
    
    with torch.no_grad():
        batch_features = model(valid_tensors)
    
    features_split = batch_features.split(1, 0)

    j = 0
    for i, tensor in enumerate(batch_tensors):
        if tensor is not None:
            feature_vectors[i] = features_split[j].squeeze(0)
            j += 1
    
    return feature_vectors

def process_rows(rows):
    image_paths = []
    bboxes = []
    bounding_box_ids = []

    for i, row in enumerate(rows):
        bounding_box_id, image_id, x, y, width, height = row
        filename = os.path.join(FOLDER_DIRECTORY, f"{image_id}.png")
        if not os.path.exists(filename) or width <= 0 or height <= 0:
            print(f"Skipping bounding_box_id {bounding_box_id}.")
            image_paths.append(None)
            continue
        
        image_paths.append(filename)
        bboxes.append((x, y, x + width, y + height))
        bounding_box_ids.append(bounding_box_id)
        ws.cell(row=i + 2, column=1, value=bounding_box_id)
        ws.cell(row=1, column=i + 2, value=bounding_box_id)
    
    feature_vectors = batch_extract_features(image_paths, bboxes)
    print(len(feature_vectors))
    print("len(rows)",len(rows))

    for i in range(len(rows)):
        for j in range(len(rows)):
            if i != j and feature_vectors[i] is not None and feature_vectors[j] is not None:
                sim = cosine(feature_vectors[i], feature_vectors[j])
                ws.cell(row=i + 2, column=j + 2, value=sim)
            else:
                ws.cell(row=i + 2, column=j + 2, value=1)

# Initialize database connection
connection = pymssql.connect(
            server="128.46.81.40",
            user="forbeslab",
            password="r^%o*3zMzsgzSxADi9",
            database="Fathom"
)
cursor = connection.cursor()

offset = 150000  # Start from the beginning or adjust as needed
query_template = "SELECT id, image_id, x, y, width, height FROM bounding_boxes ORDER BY id OFFSET {} ROWS FETCH NEXT {} ROWS ONLY"

while True:
    query = query_template.format(offset, BATCH_SIZE)
    cursor.execute(query)
    rows = cursor.fetchall()
    if not rows:
        break  # Exit the loop if no more data is returned

    for i in range(0, len(rows), IMAGE_BATCH_SIZE):
        batch_rows = rows[i:i + IMAGE_BATCH_SIZE]
        process_rows(batch_rows)
    
    offset += BATCH_SIZE

    # Force garbage collection to release unused memory
    gc.collect()

wb.save("./excel_file.xlsx")
